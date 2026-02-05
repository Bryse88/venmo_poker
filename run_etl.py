#!/usr/bin/env python3
"""
Venmo Email Parser for Gmail
Reads Venmo payment notification emails and adds them to Excel spreadsheet
"""

import os
import pickle
import base64
import re
import time
import argparse
import json
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
import openpyxl

# Gmail API scope
SCOPES = ['https://www.googleapis.com/auth/gmail.readonly']
EXCEL_PATH = "Poker.xlsx"
PROCESSED_IDS_FILE = 'processed_messages.json'


def load_processed_ids():
    """Load set of already processed message IDs from JSON file."""
    if os.path.exists(PROCESSED_IDS_FILE):
        with open(PROCESSED_IDS_FILE, 'r') as f:
            return set(json.load(f))
    return set()


def save_processed_ids(processed_ids):
    """Save processed message IDs to JSON file."""
    with open(PROCESSED_IDS_FILE, 'w') as f:
        json.dump(list(processed_ids), f)


def authenticate_gmail():
    """Authenticate with Gmail API"""
    creds = None

    # Token file stores user's access and refresh tokens
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)

    # If no valid credentials, let user log in
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)
            creds = flow.run_local_server(port=0)

        # Save credentials for next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    return build('gmail', 'v1', credentials=creds)


def get_venmo_emails(service, label_name='Venmo'):
    """Fetch emails from Venmo label"""
    try:
        # Get label ID for "Venmo"
        labels = service.users().labels().list(userId='me').execute()
        label_id = None

        for label in labels.get('labels', []):
            if label['name'].lower() == label_name.lower():
                label_id = label['id']
                break

        if not label_id:
            print(f"Label '{label_name}' not found in Gmail")
            return []

        # Fetch messages with Venmo label
        results = service.users().messages().list(
            userId='me',
            labelIds=[label_id],
            q='from:venmo@venmo.com'  # Filter for Venmo emails
        ).execute()

        messages = results.get('messages', [])
        return messages

    except Exception as e:
        print(f"Error fetching emails: {e}")
        return []


def extract_payment_info(email_body):
    """Extract payer name and dollar amount from Venmo email.

    Returns (name, amount) where:
    - Incoming payments (X paid you): positive amount
    - Outgoing payments (You paid X): negative amount
    """

    # Pattern for "FirstName LastName paid you $XX.XX" (incoming)
    incoming_pattern = r'([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+paid you\s+\$(\d+(?:\.\d{2})?)'

    # Pattern for "You paid FirstName LastName $XX.XX" (outgoing)
    outgoing_pattern = r'You paid\s+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s+\$(\d+(?:\.\d{2})?)'

    # Check for incoming payment first
    match = re.search(incoming_pattern, email_body)
    if match:
        payer_name = match.group(1)
        amount = float(match.group(2))
        return payer_name, amount

    # Check for outgoing payment
    match = re.search(outgoing_pattern, email_body)
    if match:
        payee_name = match.group(1)
        amount = -float(match.group(2))  # Negative for outgoing
        return payee_name, amount

    return None, None


def parse_email_content(service, message_id):
    """Get email content and extract payment info"""
    try:
        message = service.users().messages().get(
            userId='me',
            id=message_id,
            format='full'
        ).execute()

        # Extract email body
        parts = message.get('payload', {}).get('parts', [])
        email_body = ''

        # Handle different email structures
        if parts:
            for part in parts:
                if part.get('mimeType') == 'text/plain':
                    data = part.get('body', {}).get('data', '')
                    if data:
                        email_body = base64.urlsafe_b64decode(data).decode('utf-8')
                        break
        else:
            # Email body might be directly in payload
            data = message.get('payload', {}).get('body', {}).get('data', '')
            if data:
                email_body = base64.urlsafe_b64decode(data).decode('utf-8')

        # Also check snippet for simpler emails
        if not email_body:
            email_body = message.get('snippet', '')

        return extract_payment_info(email_body)

    except Exception as e:
        print(f"Error parsing email {message_id}: {e}")
        return None, None


def add_to_excel(excel_path, payments):
    """Add payment data to Excel sheet"""
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['money ']  # Note the space in sheet name

        # Find the next empty row (starting after row 3 which has headers)
        next_row = ws.max_row + 1

        # Add each payment
        for name, amount in payments:
            ws.cell(row=next_row, column=2, value=name)  # Column B: name
            ws.cell(row=next_row, column=3, value=amount)  # Column C: dollar amount (positive or negative)
            next_row += 1

        wb.save(excel_path)
        print(f"[OK] Added {len(payments)} payment(s) to Excel")

    except Exception as e:
        print(f"Error updating Excel: {e}")


def run_parser(service, excel_path, processed_ids):
    """Run a single parsing cycle. Returns number of new payments processed."""
    messages = get_venmo_emails(service)

    if not messages:
        print("No Venmo emails found")
        return 0

    # Filter out already processed messages
    new_messages = [msg for msg in messages if msg['id'] not in processed_ids]

    if not new_messages:
        print("No new emails to process")
        return 0

    print(f"[OK] Found {len(new_messages)} new email(s) to process")

    payments = []
    for msg in new_messages:
        name, amount = parse_email_content(service, msg['id'])
        if name and amount is not None:
            payments.append((name, amount))
            if amount >= 0:
                print(f"  [OK] Found: {name} paid ${amount:.2f}")
            else:
                print(f"  [OK] Found: You paid {name} ${-amount:.2f}")
            processed_ids.add(msg['id'])
        else:
            print(f"  [!] Could not parse message {msg['id']}")
            # Still mark as processed to avoid retrying unparseable emails
            processed_ids.add(msg['id'])

    if payments:
        add_to_excel(excel_path, payments)
    else:
        print("[!] No payment information could be extracted from new emails")

    return len(payments)


def main():
    """Main function"""
    parser = argparse.ArgumentParser(description='Venmo Email Parser for Gmail')
    parser.add_argument('--once', action='store_true',
                        help='Run once and exit instead of continuous polling')
    args = parser.parse_args()

    print("Venmo Email Parser")
    print("=" * 50)

    excel_path = EXCEL_PATH
    if not os.path.exists(excel_path):
        print(f"Error: Excel file '{excel_path}' not found!")
        return

    print("Authenticating with Gmail...")
    service = authenticate_gmail()
    print("[OK] Authentication successful")

    # Load previously processed message IDs
    processed_ids = load_processed_ids()
    print(f"[OK] Loaded {len(processed_ids)} previously processed message ID(s)")

    if args.once:
        # Single run mode
        print("\nRunning single parse cycle...")
        run_parser(service, excel_path, processed_ids)
        save_processed_ids(processed_ids)
        print("\n" + "=" * 50)
        print("[OK] Complete! Check your Excel file.")
    else:
        # Continuous polling mode
        print("\nStarting continuous polling mode (Ctrl+C to stop)...")
        try:
            while True:
                print("\n" + "=" * 50)
                print("Running Venmo email parser...")
                run_parser(service, excel_path, processed_ids)
                save_processed_ids(processed_ids)
                print("\nWaiting 5 minutes before next check...")
                time.sleep(300)  # Wait 5 minutes
        except KeyboardInterrupt:
            print("\n\nStopping parser...")
            save_processed_ids(processed_ids)
            print("[OK] Processed IDs saved. Goodbye!")


if __name__ == '__main__':
    main()
